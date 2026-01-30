import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import re
import io
from datetime import datetime
from pathlib import Path
import traceback

# Page configuration
st.set_page_config(
    page_title="Belami Asset Automation",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# Beautiful custom CSS with gradients and modern design
st.markdown("""
    <style>
    /* Import Google Fonts */
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
    
    /* Global Styles */
    * {
        font-family: 'Inter', sans-serif;
    }
    
    .main {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 2rem 1rem;
    }
    
    /* Main container */
    .block-container {
        background: white;
        border-radius: 24px;
        padding: 2rem;
        box-shadow: 0 20px 60px rgba(0,0,0,0.3);
        max-width: 1200px;
        margin: 0 auto;
    }
    
    /* Header */
    .main-title {
        font-size: 2.5rem;
        font-weight: 700;
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        text-align: center;
        margin-bottom: 0.5rem;
    }
    
    .sub-title {
        text-align: center;
        color: #6b7280;
        font-size: 1.1rem;
        margin-bottom: 2rem;
    }
    
    /* Cards */
    .config-card {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        border-radius: 16px;
        padding: 1.5rem;
        color: white;
        margin-bottom: 1.5rem;
    }
    
    .upload-card {
        background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%);
        border-radius: 16px;
        padding: 1.5rem;
        color: white;
        margin-bottom: 1.5rem;
    }
    
    .process-card {
        background: linear-gradient(135deg, #4facfe 0%, #00f2fe 100%);
        border-radius: 16px;
        padding: 1.5rem;
        color: white;
        margin-bottom: 1.5rem;
    }
    
    .card-title {
        font-size: 1.3rem;
        font-weight: 600;
        margin-bottom: 1rem;
        color: white;
    }
    
    /* Input fields */
    .stTextInput input {
        border-radius: 12px;
        border: 2px solid #e5e7eb;
        padding: 0.75rem;
        font-size: 1rem;
        transition: all 0.3s;
    }
    
    .stTextInput input:focus {
        border-color: #667eea;
        box-shadow: 0 0 0 3px rgba(102, 126, 234, 0.1);
    }
    
    .stSelectbox select {
        border-radius: 12px;
        border: 2px solid #e5e7eb;
        padding: 0.75rem;
        font-size: 1rem;
    }
    
    /* Buttons */
    .stButton > button {
        width: 100%;
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        font-weight: 600;
        padding: 1rem 2rem;
        border-radius: 12px;
        border: none;
        font-size: 1.1rem;
        transition: all 0.3s;
        box-shadow: 0 4px 15px rgba(102, 126, 234, 0.4);
    }
    
    .stButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 6px 20px rgba(102, 126, 234, 0.6);
    }
    
    .stDownloadButton > button {
        width: 100%;
        background: linear-gradient(135deg, #11998e 0%, #38ef7d 100%);
        color: white;
        font-weight: 600;
        padding: 0.75rem 1.5rem;
        border-radius: 12px;
        border: none;
        transition: all 0.3s;
    }
    
    .stDownloadButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 6px 20px rgba(17, 153, 142, 0.4);
    }
    
    /* File uploader */
    .uploadedFile {
        border-radius: 12px;
        border: 2px dashed #667eea;
        padding: 1rem;
        background: #f9fafb;
    }
    
    /* Status badges */
    .status-badge {
        display: inline-block;
        padding: 0.5rem 1rem;
        border-radius: 20px;
        font-weight: 600;
        font-size: 0.9rem;
        margin: 0.25rem;
    }
    
    .status-success {
        background: #d1fae5;
        color: #065f46;
    }
    
    .status-warning {
        background: #fef3c7;
        color: #92400e;
    }
    
    .status-info {
        background: #dbeafe;
        color: #1e40af;
    }
    
    /* Metrics */
    .metric-container {
        background: linear-gradient(135deg, #f6f8fb 0%, #ffffff 100%);
        border-radius: 12px;
        padding: 1.5rem;
        text-align: center;
        border: 2px solid #e5e7eb;
    }
    
    .metric-value {
        font-size: 2.5rem;
        font-weight: 700;
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
    }
    
    .metric-label {
        color: #6b7280;
        font-size: 0.9rem;
        margin-top: 0.5rem;
    }
    
    /* Expander */
    .streamlit-expanderHeader {
        background: #f9fafb;
        border-radius: 12px;
        font-weight: 600;
    }
    
    /* Dataframe */
    .dataframe {
        border-radius: 12px;
        overflow: hidden;
    }
    
    /* Success/Error messages */
    .stSuccess {
        background: linear-gradient(135deg, #d1fae5 0%, #a7f3d0 100%);
        border-radius: 12px;
        border-left: 4px solid #10b981;
    }
    
    .stError {
        background: linear-gradient(135deg, #fee2e2 0%, #fecaca 100%);
        border-radius: 12px;
        border-left: 4px solid #ef4444;
    }
    
    .stWarning {
        background: linear-gradient(135deg, #fef3c7 0%, #fde68a 100%);
        border-radius: 12px;
        border-left: 4px solid #f59e0b;
    }
    
    .stInfo {
        background: linear-gradient(135deg, #dbeafe 0%, #bfdbfe 100%);
        border-radius: 12px;
        border-left: 4px solid #3b82f6;
    }
    
    /* Mobile responsive */
    @media (max-width: 768px) {
        .main-title {
            font-size: 1.8rem;
        }
        .block-container {
            padding: 1rem;
        }
        .config-card, .upload-card, .process-card {
            padding: 1rem;
        }
    }
    
    /* Hide Streamlit branding */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    </style>
    """, unsafe_allow_html=True)

# Load manufacturer mapping
@st.cache_data
def load_manufacturer_mapping():
    """Load manufacturer ID mapping from uploaded file"""
    try:
        df = pd.read_excel('Manufacturer_ID_s.xlsx')
        mapping = dict(zip(df['Brand'].str.lower().str.strip(), df['Manu ID'].astype(str)))
        return mapping, df
    except Exception as e:
        st.error(f"Error loading manufacturer mapping: {e}")
        return {}, pd.DataFrame()

def clean_filename(filename):
    """Clean filename according to Belami rules"""
    name = Path(filename).stem
    name = re.sub(r'[,\s/\\]+', '_', name)
    name = re.sub(r'[^a-zA-Z0-9_-]', '', name)
    name = name.lower()
    return name

def determine_mediatype(filename, row_data):
    """Determine mediatype based on filename and context"""
    filename_lower = filename.lower()
    
    if any(word in filename_lower for word in ['lifestyle', 'room', 'environment', 'scene']):
        return 'lifestyle'
    elif any(word in filename_lower for word in ['dimension', 'measurement', 'drawing', 'diagram']):
        return 'dimension'
    elif any(word in filename_lower for word in ['detail', 'closeup', 'close-up', 'zoom']):
        return 'detail'
    elif any(word in filename_lower for word in ['angle', 'view', 'perspective']):
        return 'angle'
    elif any(word in filename_lower for word in ['swatch', 'color', 'finish']):
        return 'swatch'
    elif any(word in filename_lower for word in ['chart', 'info', 'callout', 'infographic']):
        return 'informational'
    else:
        return 'detail'

def get_brand_folder(vendor_name):
    """Get brand folder name (lowercase)"""
    brand_mapping = {
        'afx': 'afx',
        'dainolite': 'dainolite',
        'hinkley': 'hinkley',
        'justice design': 'justicedesign',
        'justicedesign': 'justicedesign',
        'crystorama': 'crystorama',
        'hudson valley': 'hudsonvalley',
    }
    
    vendor_lower = vendor_name.lower().strip()
    return brand_mapping.get(vendor_lower, vendor_lower.replace(' ', ''))

def find_columns(df, patterns, column_type):
    """Robust column finder that checks multiple patterns"""
    found_cols = []
    
    for col in df.columns:
        col_lower = str(col).lower().strip()
        for pattern in patterns:
            if pattern in col_lower:
                found_cols.append(col)
                break
    
    return found_cols

def process_vendor_file(vendor_df, mfg_prefix, brand_folder, vendor_name):
    """Process vendor file and create filled asset template"""
    
    try:
        # Find image columns
        image_patterns = [
            'image', 'photo', 'picture', 'pic', 'img',
            'url', 'link', 'path', 'file',
            'media', 'asset', 'visual',
            'main', 'primary', 'additional', 'alt',
            'http', 'www', '.jpg', '.png', '.pdf'
        ]
        
        image_cols = find_columns(vendor_df, image_patterns, "image")
        
        if not image_cols:
            return None, None, "❌ No image columns detected. Please check your file."
        
        st.success(f"✓ Found {len(image_cols)} image column(s)")
        
        # Find SKU columns
        sku_patterns = [
            'sku', 'item', 'product', 'part', 'model',
            'code', 'number', 'id', 'identifier',
            'catalog', 'reference', 'style'
        ]
        
        sku_cols = find_columns(vendor_df, sku_patterns, "SKU")
        
        if not sku_cols:
            return None, None, "❌ No SKU columns detected. Please check your file."
        
        st.success(f"✓ Found SKU column: {sku_cols[0]}")
        
        # Initialize outputs
        output_rows = []
        flags = []
        
        # Statistics
        processed_skus = 0
        total_assets = 0
        skipped_videos = 0
        main_image_count = 0
        
        # Process each row
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        total_rows = len(vendor_df)
        
        for idx, row in vendor_df.iterrows():
            progress_bar.progress((idx + 1) / total_rows)
            status_text.text(f"Processing row {idx + 1} of {total_rows}...")
            
            # Get SKU
            sku = None
            for sku_col in sku_cols:
                if pd.notna(row[sku_col]):
                    sku = str(row[sku_col]).strip()
                    if sku and sku.lower() not in ['nan', 'none', '', 'null']:
                        break
            
            if not sku:
                continue
            
            processed_skus += 1
            
            # Track images for this SKU
            images_processed = []
            is_first_image = True
            
            # Process all image columns
            for img_col in image_cols:
                if img_col not in row.index:
                    continue
                    
                img_value = row[img_col]
                
                if pd.isna(img_value) or str(img_value).strip() == '':
                    continue
                
                # Handle multiple images
                img_urls = re.split(r'[,;\n\r\|]+', str(img_value))
                
                for img_url in img_urls:
                    img_url = img_url.strip()
                    
                    if not img_url or img_url.lower() in ['nan', 'none', 'null', '']:
                        continue
                    
                    # Skip videos
                    video_extensions = ['.mp4', '.mov', '.avi', '.wmv', '.flv', '.webm', '.mkv']
                    video_sites = ['youtube', 'vimeo', 'youtu.be']
                    
                    is_video = False
                    img_lower = img_url.lower()
                    
                    if any(ext in img_lower for ext in video_extensions):
                        is_video = True
                    if any(site in img_lower for site in video_sites):
                        is_video = True
                    
                    if is_video:
                        skipped_videos += 1
                        continue
                    
                    # Extract filename
                    filename = img_url.split('/')[-1].split('\\')[-1]
                    if '?' in filename:
                        filename = filename.split('?')[0]
                    
                    if '.' not in filename:
                        filename = filename + '.jpg'
                    
                    # Check if PDF
                    is_pdf = filename.lower().endswith('.pdf')
                    
                    # Clean filename
                    clean_name = clean_filename(filename)
                    
                    if not clean_name:
                        continue
                    
                    # Determine asset type
                    if is_pdf:
                        if 'install' in filename.lower():
                            asset_family = 'install_sheet'
                        else:
                            asset_family = 'spec_sheet'
                        suffix = '_specs'
                        path_folder = 'specsheets'
                        extension = '.pdf'
                        mediatype_value = ''
                    else:
                        suffix = '_new_1k'
                        extension = '.jpg'
                        
                        if is_first_image:
                            asset_family = 'main_product_image'
                            path_folder = 'products'
                            mediatype_value = ''
                            is_first_image = False
                            main_image_count += 1
                        else:
                            asset_family = 'media'
                            path_folder = 'media'
                            mediatype_value = determine_mediatype(filename, row)
                    
                    # Create code and label
                    code_label = f"{mfg_prefix}_{clean_name}{suffix}"
                    product_ref = f"{mfg_prefix}_{sku}"
                    imagelink = f"{brand_folder}/{path_folder}/{clean_name}_new{extension}".lower()
                    
                    # Check duplicates
                    if code_label in images_processed:
                        flags.append(f"⚠️ Duplicate: {code_label}")
                    else:
                        images_processed.append(code_label)
                    
                    # Add row
                    output_rows.append({
                        'code': code_label,
                        'label-en_US': code_label,
                        'product_reference': product_ref,
                        'imagelink': imagelink,
                        'assetFamilyIdentifier': asset_family,
                        'mediatype': mediatype_value
                    })
                    
                    total_assets += 1
        
        progress_bar.empty()
        status_text.empty()
        
        # Create output DataFrame
        if not output_rows:
            return None, None, "No assets were processed."
        
        output_df = pd.DataFrame(output_rows)
        
        # Add statistics
        flags.insert(0, f"✓ Processed {processed_skus} SKUs")
        flags.insert(1, f"✓ Created {total_assets} assets")
        flags.insert(2, f"✓ Main images: {main_image_count}")
        flags.insert(3, f"✓ Videos skipped: {skipped_videos}")
        
        if main_image_count == 0:
            flags.append("⚠️ No main product images found")
        
        flags_text = "\n".join(flags)
        
        return output_df, flags_text, None
        
    except Exception as e:
        error_trace = traceback.format_exc()
        return None, None, f"Error: {str(e)}\n\n{error_trace}"

# Main UI
st.markdown('<div class="main-title">⚡ Belami Asset Automation</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-title">Transform vendor data into perfect asset templates</div>', unsafe_allow_html=True)

# Load manufacturer mapping
mfg_mapping, mfg_df = load_manufacturer_mapping()

# Initialize session state
if 'sheet_names' not in st.session_state:
    st.session_state.sheet_names = []
if 'selected_sheet' not in st.session_state:
    st.session_state.selected_sheet = None
if 'vendor_file_uploaded' not in st.session_state:
    st.session_state.vendor_file_uploaded = False

# Configuration Card
st.markdown('<div class="config-card"><div class="card-title">⚙️ Configuration</div></div>', unsafe_allow_html=True)

col1, col2, col3 = st.columns(3)

with col1:
    vendor_name = st.text_input("Vendor Name", placeholder="e.g., AFX, Dainolite", help="Enter vendor/brand name")

with col2:
    suggested_prefix = None
    if vendor_name:
        vendor_lower = vendor_name.lower().strip()
        if vendor_lower in mfg_mapping:
            suggested_prefix = mfg_mapping[vendor_lower]
    
    mfg_prefix = st.text_input(
        "Manufacturer ID", 
        value=suggested_prefix if suggested_prefix else "",
        placeholder="e.g., 2605",
        help="Manufacturer prefix code"
    )

with col3:
    brand_folder = st.text_input(
        "Brand Folder",
        value=get_brand_folder(vendor_name) if vendor_name else "",
        placeholder="e.g., afx",
        help="Lowercase brand folder name"
    )

# Upload Card
st.markdown('<div class="upload-card"><div class="card-title">📤 Upload Files</div></div>', unsafe_allow_html=True)

col1, col2 = st.columns(2)

with col1:
    vendor_file = st.file_uploader(
        "Vendor Data File",
        type=['xlsx', 'xls'],
        help="Excel file with product data",
        key="vendor_upload"
    )
    
    # Detect sheets when file is uploaded
    if vendor_file and not st.session_state.vendor_file_uploaded:
        try:
            excel_file = pd.ExcelFile(vendor_file)
            st.session_state.sheet_names = excel_file.sheet_names
            st.session_state.vendor_file_uploaded = True
            if len(st.session_state.sheet_names) == 1:
                st.session_state.selected_sheet = st.session_state.sheet_names[0]
        except Exception as e:
            st.error(f"Error reading file: {e}")
    
    # Reset when file is removed
    if not vendor_file:
        st.session_state.vendor_file_uploaded = False
        st.session_state.sheet_names = []
        st.session_state.selected_sheet = None

with col2:
    template_file = st.file_uploader(
        "Asset Template",
        type=['xlsx'],
        help="Empty Belami template",
        key="template_upload"
    )

# Sheet Selection (only show if multiple sheets)
if st.session_state.sheet_names and len(st.session_state.sheet_names) > 1:
    st.markdown("---")
    st.markdown("### 📋 Select Data Sheet")
    
    # Create buttons for each sheet
    cols = st.columns(min(len(st.session_state.sheet_names), 4))
    
    for idx, sheet_name in enumerate(st.session_state.sheet_names):
        with cols[idx % 4]:
            if st.button(
                sheet_name, 
                key=f"sheet_{idx}",
                use_container_width=True,
                type="primary" if st.session_state.selected_sheet == sheet_name else "secondary"
            ):
                st.session_state.selected_sheet = sheet_name
                st.rerun()
    
    if st.session_state.selected_sheet:
        st.success(f"✓ Selected sheet: **{st.session_state.selected_sheet}**")

# Status indicators
st.markdown("---")
col1, col2, col3, col4 = st.columns(4)

with col1:
    if vendor_name and mfg_prefix and brand_folder:
        st.markdown('<span class="status-badge status-success">✓ Config Ready</span>', unsafe_allow_html=True)
    else:
        st.markdown('<span class="status-badge status-warning">⚠ Config Needed</span>', unsafe_allow_html=True)

with col2:
    if vendor_file:
        st.markdown('<span class="status-badge status-success">✓ Vendor File</span>', unsafe_allow_html=True)
    else:
        st.markdown('<span class="status-badge status-warning">⚠ Upload File</span>', unsafe_allow_html=True)

with col3:
    if st.session_state.selected_sheet or (st.session_state.sheet_names and len(st.session_state.sheet_names) == 1):
        st.markdown('<span class="status-badge status-success">✓ Sheet Selected</span>', unsafe_allow_html=True)
    elif st.session_state.sheet_names:
        st.markdown('<span class="status-badge status-warning">⚠ Select Sheet</span>', unsafe_allow_html=True)
    else:
        st.markdown('<span class="status-badge status-info">○ Sheet Auto</span>', unsafe_allow_html=True)

with col4:
    if template_file:
        st.markdown('<span class="status-badge status-success">✓ Template</span>', unsafe_allow_html=True)
    else:
        st.markdown('<span class="status-badge status-warning">⚠ Upload Template</span>', unsafe_allow_html=True)

# Process button
st.markdown("---")
st.markdown('<div class="process-card"><div class="card-title">🚀 Generate Template</div></div>', unsafe_allow_html=True)

can_process = (vendor_file and template_file and vendor_name and mfg_prefix and brand_folder and 
               (st.session_state.selected_sheet or (st.session_state.sheet_names and len(st.session_state.sheet_names) == 1)))

if st.button("⚡ Generate Asset Template", disabled=not can_process, use_container_width=True):
    with st.spinner("Processing..."):
        try:
            # Read the selected sheet
            sheet_to_use = st.session_state.selected_sheet if st.session_state.selected_sheet else st.session_state.sheet_names[0]
            
            st.info(f"Reading sheet: **{sheet_to_use}**")
            vendor_df = pd.read_excel(vendor_file, sheet_name=sheet_to_use)
            
            st.info(f"Found **{len(vendor_df)}** rows and **{len(vendor_df.columns)}** columns")
            
            # Process the file
            output_df, flags_text, error = process_vendor_file(
                vendor_df, mfg_prefix, brand_folder, vendor_name
            )
            
            if error:
                st.error(error)
            else:
                st.balloons()
                st.success(f"🎉 Successfully created **{len(output_df)}** assets!")
                
                # Metrics
                st.markdown("---")
                st.markdown("### 📊 Results")
                
                col1, col2, col3, col4 = st.columns(4)
                
                with col1:
                    st.markdown(f'<div class="metric-container"><div class="metric-value">{len(output_df)}</div><div class="metric-label">Total Assets</div></div>', unsafe_allow_html=True)
                
                with col2:
                    main_images = len(output_df[output_df['assetFamilyIdentifier'] == 'main_product_image'])
                    st.markdown(f'<div class="metric-container"><div class="metric-value">{main_images}</div><div class="metric-label">Main Images</div></div>', unsafe_allow_html=True)
                
                with col3:
                    media_images = len(output_df[output_df['assetFamilyIdentifier'] == 'media'])
                    st.markdown(f'<div class="metric-container"><div class="metric-value">{media_images}</div><div class="metric-label">Media</div></div>', unsafe_allow_html=True)
                
                with col4:
                    pdfs = len(output_df[output_df['assetFamilyIdentifier'].isin(['spec_sheet', 'install_sheet'])])
                    st.markdown(f'<div class="metric-container"><div class="metric-value">{pdfs}</div><div class="metric-label">PDFs</div></div>', unsafe_allow_html=True)
                
                # Preview
                st.markdown("---")
                with st.expander("📋 Preview Data", expanded=True):
                    st.dataframe(output_df.head(20), use_container_width=True)
                
                # Download buttons
                st.markdown("---")
                st.markdown("### 💾 Download")
                
                col1, col2 = st.columns(2)
                
                with col1:
                    output_excel = io.BytesIO()
                    with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
                        output_df.to_excel(writer, sheet_name='Sheet1', index=False)
                    output_excel.seek(0)
                    
                    st.download_button(
                        label="📥 Download Asset Template",
                        data=output_excel,
                        file_name=f"{vendor_name}_Asset_Template.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                
                with col2:
                    flags_file = f"Processing Log - {vendor_name}\n"
                    flags_file += f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
                    flags_file += "=" * 50 + "\n\n"
                    flags_file += flags_text
                    
                    st.download_button(
                        label="📄 Download Log",
                        data=flags_file,
                        file_name=f"{vendor_name}_log.txt",
                        mime="text/plain",
                        use_container_width=True
                    )
        
        except Exception as e:
            st.error(f"Error: {str(e)}")
            st.code(traceback.format_exc())

# Footer
st.markdown("---")
st.markdown(
    '<div style="text-align: center; color: #9ca3af; padding: 1rem;">Belami Asset Automation v2.0</div>', 
    unsafe_allow_html=True
)
import pandas as pd
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import re
import io
from datetime import datetime
from pathlib import Path
import traceback

# Page configuration
st.set_page_config(
    page_title="Belami Asset Template Automation",
    page_icon="📋",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for professional styling
st.markdown("""
    <style>
    .main-header {
        font-size: 2.5rem;
        font-weight: 600;
        color: #1f2937;
        margin-bottom: 0.5rem;
    }
    .sub-header {
        font-size: 1.1rem;
        color: #6b7280;
        margin-bottom: 2rem;
    }
    .section-header {
        font-size: 1.3rem;
        font-weight: 600;
        color: #374151;
        margin-top: 1.5rem;
        margin-bottom: 1rem;
        border-bottom: 2px solid #e5e7eb;
        padding-bottom: 0.5rem;
    }
    .info-box {
        background-color: #f3f4f6;
        padding: 1rem;
        border-radius: 0.5rem;
        border-left: 4px solid #3b82f6;
        margin: 1rem 0;
    }
    .success-box {
        background-color: #d1fae5;
        padding: 1rem;
        border-radius: 0.5rem;
        border-left: 4px solid #10b981;
        margin: 1rem 0;
    }
    .warning-box {
        background-color: #fef3c7;
        padding: 1rem;
        border-radius: 0.5rem;
        border-left: 4px solid #f59e0b;
        margin: 1rem 0;
    }
    .stButton>button {
        width: 100%;
        background-color: #3b82f6;
        color: white;
        font-weight: 500;
        padding: 0.5rem 1rem;
        border-radius: 0.375rem;
        border: none;
    }
    .stButton>button:hover {
        background-color: #2563eb;
    }
    </style>
    """, unsafe_allow_html=True)

# Load manufacturer mapping
@st.cache_data
def load_manufacturer_mapping():
    """Load manufacturer ID mapping from uploaded file"""
    try:
        df = pd.read_excel('Manufacturer_ID_s.xlsx')
        # Create mapping dictionary (brand name -> ID)
        mapping = dict(zip(df['Brand'].str.lower().str.strip(), df['Manu ID'].astype(str)))
        return mapping, df
    except Exception as e:
        st.error(f"Error loading manufacturer mapping: {e}")
        return {}, pd.DataFrame()

def clean_filename(filename):
    """Clean filename according to Belami rules"""
    # Remove extension
    name = Path(filename).stem
    # Replace spaces, commas, slashes, special chars with underscore
    name = re.sub(r'[,\s/\\]+', '_', name)
    # Remove other special characters except underscore and hyphen
    name = re.sub(r'[^a-zA-Z0-9_-]', '', name)
    # Make lowercase
    name = name.lower()
    return name

def determine_mediatype(filename, row_data):
    """Determine mediatype based on filename and context"""
    filename_lower = filename.lower()
    
    # Keywords for different media types
    if any(word in filename_lower for word in ['lifestyle', 'room', 'environment', 'scene']):
        return 'lifestyle'
    elif any(word in filename_lower for word in ['dimension', 'measurement', 'drawing', 'diagram']):
        return 'dimension'
    elif any(word in filename_lower for word in ['detail', 'closeup', 'close-up', 'zoom']):
        return 'detail'
    elif any(word in filename_lower for word in ['angle', 'view', 'perspective']):
        return 'angle'
    elif any(word in filename_lower for word in ['swatch', 'color', 'finish']):
        return 'swatch'
    elif any(word in filename_lower for word in ['chart', 'info', 'callout', 'infographic']):
        return 'informational'
    else:
        # Default to detail if unclear
        return 'detail'

def get_brand_folder(vendor_name):
    """Get brand folder name (lowercase)"""
    # Map common vendor names to brand folders
    brand_mapping = {
        'afx': 'afx',
        'dainolite': 'dainolite',
        'hinkley': 'hinkley',
        'justice design': 'justicedesign',
        'justicedesign': 'justicedesign',
        'crystorama': 'crystorama',
        'hudson valley': 'hudsonvalley',
    }
    
    vendor_lower = vendor_name.lower().strip()
    return brand_mapping.get(vendor_lower, vendor_lower.replace(' ', ''))

def find_columns(df, patterns, column_type):
    """Robust column finder that checks multiple patterns"""
    found_cols = []
    
    for col in df.columns:
        col_lower = str(col).lower().strip()
        for pattern in patterns:
            if pattern in col_lower:
                found_cols.append(col)
                break
    
    if not found_cols:
        st.warning(f"⚠️ No {column_type} columns found automatically. Available columns:")
        st.write(list(df.columns))
    
    return found_cols

def process_vendor_file(vendor_file, template_file, mfg_mapping, vendor_name, mfg_prefix, brand_folder):
    """Process vendor file and create filled asset template"""
    
    try:
        # Read vendor file - try multiple approaches
        vendor_df = None
        sheet_used = None
        
        # Get all sheet names first
        try:
            excel_file = pd.ExcelFile(vendor_file)
            available_sheets = excel_file.sheet_names
            st.info(f"📋 Available sheets: {', '.join(available_sheets)}")
        except Exception as e:
            return None, None, f"Error reading Excel file structure: {str(e)}"
        
        # Try common sheet names in order
        sheet_priority = ['Entities', 'entities', 'All', 'all', 'Sheet1', 'sheet1', 'Data', 'data']
        
        for sheet_name in sheet_priority:
            if sheet_name in available_sheets:
                try:
                    vendor_df = pd.read_excel(vendor_file, sheet_name=sheet_name)
                    sheet_used = sheet_name
                    st.success(f"✓ Using sheet: {sheet_name}")
                    break
                except Exception as e:
                    continue
        
        # If no priority sheet found, use first sheet
        if vendor_df is None:
            try:
                vendor_df = pd.read_excel(vendor_file, sheet_name=0)
                sheet_used = available_sheets[0]
                st.success(f"✓ Using first sheet: {sheet_used}")
            except Exception as e:
                return None, None, f"Error reading Excel file: {str(e)}"
        
        # Display column information
        st.info(f"📊 Found {len(vendor_df.columns)} columns and {len(vendor_df)} rows")
        
        with st.expander("Show all columns in vendor file"):
            st.write(list(vendor_df.columns))
        
        # Find image columns with expanded patterns
        image_patterns = [
            'image', 'photo', 'picture', 'pic', 'img',
            'url', 'link', 'path', 'file',
            'media', 'asset', 'visual',
            'main', 'primary', 'additional', 'alt',
            'http', 'www', '.jpg', '.png', '.pdf'
        ]
        
        image_cols = find_columns(vendor_df, image_patterns, "image")
        
        if not image_cols:
            # Manual column selection
            st.error("❌ Could not auto-detect image columns")
            st.write("**Please select image columns manually:**")
            return None, None, "No image columns detected. Please check your file format."
        
        st.success(f"✓ Found {len(image_cols)} image column(s): {', '.join(image_cols)}")
        
        # Find SKU columns with expanded patterns
        sku_patterns = [
            'sku', 'item', 'product', 'part', 'model',
            'code', 'number', 'id', 'identifier',
            'catalog', 'reference', 'style'
        ]
        
        sku_cols = find_columns(vendor_df, sku_patterns, "SKU")
        
        if not sku_cols:
            st.error("❌ Could not auto-detect SKU columns")
            return None, None, "No SKU columns detected. Please check your file format."
        
        st.success(f"✓ Found SKU column: {sku_cols[0]}")
        
        # Initialize outputs
        output_rows = []
        flags = []
        
        # Statistics
        total_rows = 0
        skipped_rows = 0
        processed_skus = 0
        total_assets = 0
        skipped_videos = 0
        
        # Process each row
        main_image_count = 0
        
        for idx, row in vendor_df.iterrows():
            total_rows += 1
            
            # Get SKU
            sku = None
            for sku_col in sku_cols:
                if pd.notna(row[sku_col]):
                    sku = str(row[sku_col]).strip()
                    if sku and sku.lower() not in ['nan', 'none', '', 'null']:
                        break
            
            if not sku:
                skipped_rows += 1
                continue
            
            processed_skus += 1
            
            # Track images for this SKU
            images_processed = []
            is_first_image = True
            sku_asset_count = 0
            
            # Process all image columns
            for img_col in image_cols:
                if img_col not in row.index:
                    continue
                    
                img_value = row[img_col]
                
                if pd.isna(img_value) or str(img_value).strip() == '':
                    continue
                
                # Handle multiple images in one cell (separated by various delimiters)
                img_urls = re.split(r'[,;\n\r\|]+', str(img_value))
                
                for img_url in img_urls:
                    img_url = img_url.strip()
                    
                    if not img_url or img_url.lower() in ['nan', 'none', 'null', '']:
                        continue
                    
                    # Skip videos
                    video_extensions = ['.mp4', '.mov', '.avi', '.wmv', '.flv', '.webm', '.mkv']
                    video_sites = ['youtube', 'vimeo', 'youtu.be']
                    
                    is_video = False
                    img_lower = img_url.lower()
                    
                    if any(ext in img_lower for ext in video_extensions):
                        is_video = True
                    if any(site in img_lower for site in video_sites):
                        is_video = True
                    
                    if is_video:
                        flags.append(f"⏭️ Skipped video for SKU {sku}: {img_url[:50]}...")
                        skipped_videos += 1
                        continue
                    
                    # Extract filename from URL or path
                    # Handle various URL formats
                    filename = img_url.split('/')[-1].split('\\')[-1]
                    if '?' in filename:
                        filename = filename.split('?')[0]
                    
                    # If no extension found, assume .jpg
                    if '.' not in filename:
                        filename = filename + '.jpg'
                    
                    # Check if it's a PDF
                    is_pdf = filename.lower().endswith('.pdf')
                    
                    # Clean filename
                    clean_name = clean_filename(filename)
                    
                    if not clean_name:
                        flags.append(f"⚠️ Could not clean filename for SKU {sku}: {filename}")
                        continue
                    
                    # Determine asset type
                    if is_pdf:
                        # Check if spec or install sheet
                        if 'install' in filename.lower():
                            asset_family = 'install_sheet'
                            suffix = '_specs'
                            path_folder = 'specsheets'
                            extension = '.pdf'
                        else:
                            asset_family = 'spec_sheet'
                            suffix = '_specs'
                            path_folder = 'specsheets'
                            extension = '.pdf'
                        
                        mediatype_value = ''
                        
                    else:
                        # Image file
                        suffix = '_new_1k'
                        extension = '.jpg'
                        
                        if is_first_image:
                            asset_family = 'main_product_image'
                            path_folder = 'products'
                            mediatype_value = ''
                            is_first_image = False
                            main_image_count += 1
                        else:
                            asset_family = 'media'
                            path_folder = 'media'
                            mediatype_value = determine_mediatype(filename, row)
                    
                    # Create code and label (identical and lowercase)
                    code_label = f"{mfg_prefix}_{clean_name}{suffix}"
                    
                    # Create product_reference
                    product_ref = f"{mfg_prefix}_{sku}"
                    
                    # Create imagelink (all lowercase)
                    imagelink = f"{brand_folder}/{path_folder}/{clean_name}_new{extension}".lower()
                    
                    # Check for duplicates
                    if code_label in images_processed:
                        flags.append(f"⚠️ Duplicate filename for SKU {sku}: {code_label}")
                    else:
                        images_processed.append(code_label)
                    
                    # Add row
                    output_rows.append({
                        'code': code_label,
                        'label-en_US': code_label,
                        'product_reference': product_ref,
                        'imagelink': imagelink,
                        'assetFamilyIdentifier': asset_family,
                        'mediatype': mediatype_value
                    })
                    
                    sku_asset_count += 1
                    total_assets += 1
            
            if sku_asset_count == 0:
                flags.append(f"⚠️ No assets found for SKU: {sku}")
        
        # Create output DataFrame
        if not output_rows:
            return None, None, "No assets were processed. Please check your vendor file format."
        
        output_df = pd.DataFrame(output_rows)
        
        # Add processing statistics
        flags.insert(0, f"📊 Processing Summary:")
        flags.insert(1, f"  • Total rows in file: {total_rows}")
        flags.insert(2, f"  • SKUs processed: {processed_skus}")
        flags.insert(3, f"  • Total assets created: {total_assets}")
        flags.insert(4, f"  • Main images: {main_image_count}")
        flags.insert(5, f"  • Videos skipped: {skipped_videos}")
        flags.insert(6, f"  • Rows skipped (no SKU): {skipped_rows}")
        flags.insert(7, "")
        
        # Validation checks
        if main_image_count == 0:
            flags.append("⚠️ WARNING: No main product images found")
        
        if main_image_count < processed_skus:
            flags.append(f"⚠️ WARNING: Some SKUs missing main images ({main_image_count} main images for {processed_skus} SKUs)")
        
        # Create flags text
        flags_text = "\n".join(flags) if flags else "No issues detected"
        
        return output_df, flags_text, None
        
    except Exception as e:
        error_trace = traceback.format_exc()
        st.error("❌ Error during processing:")
        st.code(error_trace)
        return None, None, f"Processing error: {str(e)}\n\nFull trace:\n{error_trace}"

# Main UI
st.markdown('<div class="main-header">Belami Asset Template Automation</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-header">Automatically fill Asset Templates with vendor data according to Belami standards</div>', unsafe_allow_html=True)

# Load manufacturer mapping
mfg_mapping, mfg_df = load_manufacturer_mapping()

# Sidebar for configuration
with st.sidebar:
    st.markdown("### Configuration")
    
    st.markdown("#### Vendor Information")
    vendor_name = st.text_input("Vendor Name", help="e.g., AFX, Dainolite, Hinkley")
    
    # Auto-suggest manufacturer prefix
    suggested_prefix = None
    if vendor_name:
        vendor_lower = vendor_name.lower().strip()
        if vendor_lower in mfg_mapping:
            suggested_prefix = mfg_mapping[vendor_lower]
            st.success(f"Found manufacturer ID: {suggested_prefix}")
    
    mfg_prefix = st.text_input(
        "Manufacturer Prefix", 
        value=suggested_prefix if suggested_prefix else "",
        help="Manufacturer ID (e.g., 2605 for AFX)"
    )
    
    brand_folder = st.text_input(
        "Brand Folder Name",
        value=get_brand_folder(vendor_name) if vendor_name else "",
        help="Lowercase brand folder (e.g., afx, dainolite)"
    )
    
    st.markdown("---")
    
    st.markdown("#### File Upload")
    st.info("Upload the required files below")

# Main content area
col1, col2 = st.columns([2, 1])

with col1:
    st.markdown('<div class="section-header">File Upload</div>', unsafe_allow_html=True)
    
    # Upload vendor file
    vendor_file = st.file_uploader(
        "Upload Vendor Data File",
        type=['xlsx', 'xls'],
        help="Excel file with vendor product data (usually 'Entities' or 'All' sheet)"
    )
    
    # Show preview if file uploaded
    if vendor_file:
        try:
            with st.expander("📋 Preview Vendor File (first 5 rows)"):
                preview_df = pd.read_excel(vendor_file, sheet_name=0, nrows=5)
                st.dataframe(preview_df, use_container_width=True)
                st.caption(f"Columns: {', '.join(list(preview_df.columns))}")
        except Exception as e:
            st.error(f"Could not preview file: {e}")
    
    # Upload template file
    template_file = st.file_uploader(
        "Upload Asset Template",
        type=['xlsx'],
        help="Empty Belami Asset Template (Asset Template.xlsx)"
    )

with col2:
    st.markdown('<div class="section-header">Status</div>', unsafe_allow_html=True)
    
    # Status indicators
    status_items = []
    if vendor_name and mfg_prefix:
        status_items.append("✓ Vendor configured")
    else:
        status_items.append("⚠ Vendor info needed")
    
    if vendor_file:
        status_items.append("✓ Vendor file uploaded")
    else:
        status_items.append("⚠ Vendor file needed")
    
    if template_file:
        status_items.append("✓ Template uploaded")
    else:
        status_items.append("⚠ Template needed")
    
    for item in status_items:
        if "✓" in item:
            st.markdown(f'<div style="color: #10b981; padding: 0.25rem 0;">{item}</div>', unsafe_allow_html=True)
        else:
            st.markdown(f'<div style="color: #f59e0b; padding: 0.25rem 0;">{item}</div>', unsafe_allow_html=True)

# Process button
st.markdown('<div class="section-header">Processing</div>', unsafe_allow_html=True)

can_process = vendor_file and template_file and vendor_name and mfg_prefix and brand_folder

if not can_process:
    st.warning("Please provide all required information and files before processing")

# Add manual column selection option
if vendor_file and not can_process:
    try:
        excel_file = pd.ExcelFile(vendor_file)
        test_df = pd.read_excel(vendor_file, sheet_name=0, nrows=5)
        
        with st.expander("🔧 Advanced: Manual Column Selection"):
            st.info("If automatic detection fails, you can manually select columns here")
            
            col1, col2 = st.columns(2)
            with col1:
                sku_column_manual = st.selectbox(
                    "Select SKU Column",
                    options=["Auto-detect"] + list(test_df.columns),
                    help="Column containing product SKUs/Item numbers"
                )
            
            with col2:
                image_columns_manual = st.multiselect(
                    "Select Image Columns",
                    options=list(test_df.columns),
                    help="Columns containing image URLs or filenames"
                )
    except:
        pass

if st.button("Generate Asset Template", disabled=not can_process, type="primary"):
    with st.spinner("Processing vendor data..."):
        # Process the file
        output_df, flags_text, error = process_vendor_file(
            vendor_file, template_file, mfg_mapping, 
            vendor_name, mfg_prefix, brand_folder
        )
        
        if error:
            st.error(f"Error: {error}")
        else:
            st.success(f"Successfully processed {len(output_df)} assets")
            
            # Display results
            st.markdown('<div class="section-header">Results</div>', unsafe_allow_html=True)
            
            # Summary metrics
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                st.metric("Total Assets", len(output_df))
            
            with col2:
                main_images = len(output_df[output_df['assetFamilyIdentifier'] == 'main_product_image'])
                st.metric("Main Images", main_images)
            
            with col3:
                media_images = len(output_df[output_df['assetFamilyIdentifier'] == 'media'])
                st.metric("Media Images", media_images)
            
            with col4:
                pdfs = len(output_df[output_df['assetFamilyIdentifier'].isin(['spec_sheet', 'install_sheet'])])
                st.metric("PDFs", pdfs)
            
            # Preview data
            st.markdown("#### Data Preview")
            st.dataframe(output_df.head(20), use_container_width=True)
            
            # Flags/warnings
            if flags_text != "No issues detected":
                st.markdown("#### Validation Flags")
                st.markdown(f'<div class="warning-box">{flags_text.replace(chr(10), "<br>")}</div>', unsafe_allow_html=True)
            
            # Download files
            st.markdown('<div class="section-header">Download Results</div>', unsafe_allow_html=True)
            
            col1, col2 = st.columns(2)
            
            with col1:
                # Create Excel file
                output_excel = io.BytesIO()
                with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
                    output_df.to_excel(writer, sheet_name='Sheet1', index=False)
                output_excel.seek(0)
                
                st.download_button(
                    label="Download Filled Asset Template",
                    data=output_excel,
                    file_name=f"{vendor_name}_Filled_Asset_Template.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            
            with col2:
                # Create flags file
                flags_file = f"Asset Template Processing Log - {vendor_name}\n"
                flags_file += f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
                flags_file += f"Total Assets: {len(output_df)}\n\n"
                flags_file += "Validation Flags:\n"
                flags_file += "=" * 50 + "\n"
                flags_file += flags_text
                
                st.download_button(
                    label="Download Processing Log",
                    data=flags_file,
                    file_name=f"{vendor_name}_asset_flags.txt",
                    mime="text/plain"
                )

# Information section at bottom
with st.expander("Processing Rules & Guidelines"):
    st.markdown("""
    ### Belami Asset Template Processing Rules
    
    **File Naming Convention:**
    - Images: `manufacturerprefix_filename_without_extension_new_1k`
    - PDFs: `manufacturerprefix_filename_without_extension_specs`
    - All filenames are converted to lowercase
    
    **Asset Family Types:**
    - `main_product_image`: First image with white background (full product view)
    - `media`: Additional product images
    - `spec_sheet`: Specification PDFs
    - `install_sheet`: Installation instruction PDFs
    
    **Media Types (for media assets only):**
    - `lifestyle`: Product in room/environment
    - `dimension`: Measurements, drawings, diagrams
    - `detail`: Close-up views
    - `angle`: Alternate viewing angles
    - `informational`: Charts, callouts, infographics
    - `swatch`: Color/finish swatches
    
    **Image Link Paths:**
    - Main images: `brand/products/filename_new_1k.jpg`
    - Media images: `brand/media/filename_new_1k.jpg`
    - PDFs: `brand/specsheets/filename_new.pdf`
    
    **Important Notes:**
    - Video files are automatically skipped
    - Each row represents one real asset (image or PDF)
    - All paths and filenames must be lowercase
    - No empty columns in generated rows
    """)

# Footer
st.markdown("---")
st.markdown(
    '<div style="text-align: center; color: #6b7280; padding: 1rem;">Belami Asset Template Automation Tool | Version 1.0</div>', 
    unsafe_allow_html=True
)
